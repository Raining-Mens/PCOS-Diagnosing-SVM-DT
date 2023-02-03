from flask import Flask, request, jsonify, redirect, url_for, render_template, session
import pickle
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import git

THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
EXCEL_UPLOADS = "static/assets/uploads"
ASSETS = "static/assets"
my_excel = os.path.join(THIS_FOLDER, "static/assets/uploads")
my_assets = os.path.join(THIS_FOLDER, "static/assets")


def predict_excel_svm(excel):
    wb = load_workbook(excel)

    ws = wb.active

    PatID = ws["A2"].value     

    #Integer Number
    Age = ws["B2"].value
    Weight = ws["C2"].value
    Height = ws["D2"].value
    Hip = ws["E2"].value
    Waist = ws["H2"].value
    WaistRatio = ws["G2"].value
    HairLoss = ws["K2"].value
    RegExer = ws["N2"].value
    Bmi = ws["O2"].value
    BloodGrp = ws["P2"].value
    PulseRate = ws["Q2"].value
    Rr = ws["R2"].value
    Hb = ws["S2"].value
    Marraige = ws["V2"].value
    Pregnant = ws["W2"].value
    Absorptions = ws["X2"].value
    betaHCG1 = ws["Y2"].value
    betaHCG2 = ws["Z2"].value
    Fsh = ws["AA2"].value
    Lh = ws["AB2"].value
    FshLH = ws["AC2"].value
    Tsh = ws["AD2"].value
    Amh = ws["AE2"].value
    Prl = ws["AF2"].value
    VitD3 = ws["AG2"].value
    Prg = ws["AH2"].value
    Rbs = ws["AI2"].value
    BPSys = ws["AJ2"].value
    BPDias = ws["AK2"].value
    FoliL = ws["AL2"].value
    FoliR = ws["AM2"].value
    Endo = ws["AP2"].value
    CycleLength = ws["U2"].value

    # 1 OR 0
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    WeightGain = ws["H2"].value
    FastFood = ws["M2"].value
    Pimple = ws["L2"].value

    # 2 OR 4
    CycleRI = ws["T2"].value

    # Int or Float
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value

    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["WeightGain"] = WeightGain
    session["FastFood"] = FastFood


    model = pickle.load(open(os.path.join(my_assets, "model-svm.pkl"), 'rb'))
    session['model'] = "SVM"


    makeprediction = model.predict([[Hairgrowth, SkinDarkening,
                                    WeightGain,  
                                    CycleLength, AvgFsizeLmm, AvgFsizeRmm,
                                    FastFood, Pimple, CycleRI
                                    ]])

    output = round(makeprediction[0])

    return(output)

def predict_excel_dt(excel):
    wb = load_workbook(excel)

    ws = wb.active

    
    PatID = ws["A2"].value

    #Integer Number
    Age = ws["B2"].value
    CycleLength = ws["U2"].value

    if not(isinstance(Age, int) and isinstance(CycleLength, int)):
        session["errorint"] = "Incorrect Data"

    # 1 OR 0
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    WeightGain = ws["H2"].value
    FastFood = ws["M2"].value
    Pimple = ws["L2"].value

    # 2 OR 4
    CycleRI = ws["T2"].value

    # Int or Float
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value

    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["WeightGain"] = WeightGain
    session["FastFood"] = FastFood


    model = pickle.load(open(os.path.join(my_assets, "model-svm.pkl"), 'rb'))
    session['model'] = "DT"


    makeprediction = model.predict([[Hairgrowth, SkinDarkening,
                                    WeightGain, CycleRI, FastFood, Pimple,
                                    CycleLength, AvgFsizeLmm, AvgFsizeRmm]])

    output = round(makeprediction[0], 2)

    return(output)


def ovarian_svm(excel):
    wb = load_workbook(excel)

    ws = wb.active

    Age = ws["C2"].value
    Menopause = ws["AH2"].value
    CANine = ws["M2"].value
    CASeven = ws["N2"].value
    AeFP = ws["A2"].value
    CAOneTwo = ws["L2"].value
    HEFour = ws["Z2"].value
    CEyA = ws["O2"].value


    session["Age"] = Age
    session["Menopause"] = Menopause
    session["CANine"] = CANine
    session["CASeven"] = CASeven
    session["AeFP"] = AeFP
    session["CAOneTwo"] = CAOneTwo


    model = pickle.load(open(os.path.join(my_assets, "svm-ovarian.pkl"), 'rb'))
    session['model'] = "SVM"


    makeprediction = model.predict([[Age, Menopause, CANine,
                                    CASeven, AeFP, CAOneTwo, HEFour,
                                    CEyA]])

    output = round(makeprediction[0], 2)

    return(output)


def ovarian_dt(excel):
    wb = load_workbook(excel)

    ws = wb.active

    Age = ws["C2"].value
    Menopause = ws["AH2"].value
    CANine = ws["M2"].value
    CASeven = ws["N2"].value
    AeFP = ws["A2"].value
    CAOneTwo = ws["L2"].value
    HEFour = ws["Z2"].value
    CEyA = ws["O2"].value


    session["Age"] = Age
    session["Menopause"] = Menopause
    session["CANine"] = CANine
    session["CASeven"] = CASeven
    session["AeFP"] = AeFP
    session["CAOneTwo"] = CAOneTwo


    model = pickle.load(open(os.path.join(my_assets, "dt-ovarian.pkl"), 'rb'))
    session['model'] = "DT"


    makeprediction = model.predict([[Age, Menopause, CANine,
                                    CASeven, AeFP, CAOneTwo, HEFour,
                                    CEyA]])

    output = round(makeprediction[0], 2)

    return(output)