# Separate function file of predicting the data from the excel then producing the results

# Importing modules
from flask import Flask, request, jsonify, redirect, url_for, render_template, session
import pickle
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os

# Configurations
THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
EXCEL_UPLOADS = "static/assets/uploads"
ASSETS = "static/assets"
my_excel = os.path.join(THIS_FOLDER, "static/assets/uploads")
my_assets = os.path.join(THIS_FOLDER, "static/assets")

# Function to predict the PCOS SVM classifier
def predict_pcos_svm(excel):
    wb = load_workbook(excel) # Loading the excel file
    ws = wb.active

    # Reading the excel file's cell value and assigning a variable
    PatID = ws["A2"].value
    Age = ws["B2"].value
    CycleLength = ws["U2"].value
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    WeightGain = ws["H2"].value
    FastFood = ws["M2"].value
    Pimple = ws["L2"].value
    CycleRI = ws["T2"].value
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value

    # Storing the values to a session to be displayed on the results page
    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["WeightGain"] = WeightGain
    session["FastFood"] = FastFood

    # Loading the machine learning model to be used
    model = pickle.load(open(os.path.join(my_assets, "svm-model.pkl"), 'rb'))
    session['model'] = "SVM" # Storing the model's name to a session

    # Prediction function using the variables from the cell values
    prediction = model.predict([[Hairgrowth, SkinDarkening,
                                 WeightGain, CycleLength,
                                 AvgFsizeLmm, AvgFsizeRmm,
                                 FastFood, Pimple, CycleRI
                                 ]])

    output = round(prediction[0]) # Rounding the prediction from float to a whole number then storing it to 'output'
    return(output)

# Function to predict the PCOS Decision Tree classifier
def predict_pcos_dt(excel):
    wb = load_workbook(excel)
    ws = wb.active

    PatID = ws["A2"].value
    Age = ws["B2"].value
    CycleLength = ws["U2"].value
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    WeightGain = ws["H2"].value
    FastFood = ws["M2"].value
    Pimple = ws["L2"].value
    CycleRI = ws["T2"].value
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value

    session["PatID"] = PatID
    session["Age"] = Age
    session["Hairgrowth"] = Hairgrowth
    session["CycleRI"] = CycleRI
    session["WeightGain"] = WeightGain
    session["FastFood"] = FastFood

    model = pickle.load(open(os.path.join(my_assets, "dt-model.pkl"), 'rb'))
    session['model'] = "DT"

    prediction = model.predict([[Hairgrowth, SkinDarkening,
                                 WeightGain, CycleLength,
                                 AvgFsizeLmm, AvgFsizeRmm,
                                 FastFood, Pimple, CycleRI
                                 ]])

    output = round(prediction[0])
    return(output)

# Function to predict the Ovarian Cancer SVM classifier
def predict_ovarian_svm(excel):
    wb = load_workbook(excel)
    ws = wb.active

    Age = ws["C2"].value
    Menopause = ws["AH2"].value
    CANine = ws["M2"].value
    CASeven = ws["N2"].value
    AeFP = ws["A2"].value
    CAOneTwo = ws["L2"].value
    HEFour = ws["Z2"].value
    CEyA = ws["O2"].value

    session["Age"] = Age
    session["Menopause"] = Menopause
    session["CANine"] = CANine
    session["CASeven"] = CASeven
    session["AeFP"] = AeFP
    session["CAOneTwo"] = CAOneTwo

    model = pickle.load(open(os.path.join(my_assets, "svm-ovarian.pkl"), 'rb'))
    session['model'] = "SVM"

    prediction = model.predict([[Age, Menopause, CANine,
                                 CASeven, AeFP, CAOneTwo,
                                 HEFour, CEyA
                                 ]])

    output = round(prediction[0])
    return(output)

# Function to predict the Ovarian Cancer Decision Tree classifier
def predict_ovarian_dt(excel):
    wb = load_workbook(excel)
    ws = wb.active

    Age = ws["C2"].value
    Menopause = ws["AH2"].value
    CANine = ws["M2"].value
    CASeven = ws["N2"].value
    AeFP = ws["A2"].value
    CAOneTwo = ws["L2"].value
    HEFour = ws["Z2"].value
    CEyA = ws["O2"].value

    session["Age"] = Age
    session["Menopause"] = Menopause
    session["CANine"] = CANine
    session["CASeven"] = CASeven
    session["AeFP"] = AeFP
    session["CAOneTwo"] = CAOneTwo

    model = pickle.load(open(os.path.join(my_assets, "dt-ovarian.pkl"), 'rb'))
    session['model'] = "DT"

    prediction = model.predict([[Age, Menopause, CANine,
                                 CASeven, AeFP, CAOneTwo,
                                 HEFour, CEyA
                                 ]])

    output = round(prediction[0])
    return(output)
