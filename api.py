from flask import Flask, request, jsonify, redirect, url_for, render_template, session
import pickle
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

app = Flask(__name__)

app.secret_key = "hello"

@app.route("/")
def home():
    return render_template("index.html")

app.config["EXCEL_UPLOADS"] = "./static/assets/uploads"
app.config["ALLOWED_EXCEL_EXTENSIONS"] = ["XLSX", "CSV", "XLS"]

def predict_excel(excel):
    wb = load_workbook(excel)

    ws = wb.active

    Age = ws["B2"].value
    Hairgrowth = ws["I2"].value
    SkinDarkening = ws["J2"].value
    PulseRateBPM = ws["Q2"].value
    CycleRI = ws["T2"].value
    FSHmIUmL = ws["AA2"].value
    LHmIUmL = ws["AB2"].value
    AMHngmL = ws["AE2"].value
    PRGngmL = ws["AH2"].value
    RBSmgdl = ws["AI2"].value
    BP_SystolicmmHg = ws["AJ2"].value
    BP_DiastolicmmHg = ws["AK2"].value
    AvgFsizeLmm = ws["AN2"].value
    AvgFsizeRmm = ws["AO2"].value
    Endometriummm = ws["AP2"].value

    radio = request.form['radio']
    if radio == "SVM":
        model = pickle.load(open('models\svm-model.pkl', 'rb'))
    elif radio == "DT":
        model = pickle.load(open('models\dt-model.pkl', 'rb'))
    else:
        redirect(url_for("tool"))

    makeprediction = model.predict([[Age, Hairgrowth, SkinDarkening,
                                    PulseRateBPM, CycleRI, FSHmIUmL, LHmIUmL,
                                    AMHngmL, PRGngmL, RBSmgdl, BP_SystolicmmHg,
                                    BP_DiastolicmmHg, AvgFsizeLmm, AvgFsizeRmm, Endometriummm]])

    output = round(makeprediction[0], 2)

    return(output)


@app.route("/tool", methods=["GET", "POST"])
def tool():

    if request.method == "POST":
        if request.files:

            excel = request.files["input"]
            output = predict_excel(excel)
            print(output)

            session['result'] = int(output)

            return redirect(url_for("result"))
    else:    
        return render_template("tool.html")


@app.route("/result", methods=["GET", "POST"])
def result():
    if "result" in session:
        result = session["result"]
        return f"<h1>{result}</h1>"
    else:
        return redirect(url_for("tool"))


if __name__ == "__main__":
    app.run(debug=True)
