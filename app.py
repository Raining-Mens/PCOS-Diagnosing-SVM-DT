# Tool for thesis entitled 
# "An Ensemble ADASYN-ENN Resampling Approach in Diagnosing PCOS"
# "OvYou" Web Application
# .
# Developers:
#     Colasino, Jayson Kim G.
#     Fatallo, Lance Raphael F.
#     Gatchalian, Jan Kristian B.
#     Pascua, Karl Melo F.
#
# The OvYou Web App, 'app.py' is the main application runner which contains all routes and templates that will be run by this file.
# .
# This app was started on September 2022 then finished on February 2023
# .
# Its main purpose is to diagnose a patient whether it has Polycystic Ovary Syndrome or not using the patient's data
# .
# All data are stored and retrieved through Sessions. See more on https://bit.ly/3KdUGH5


# The codebase use the Anaconda3 2022.10 (Python 3.9.13) https://bit.ly/411MAqP as the Python Interpreter and library of modules
# After installing Anaconda3, go to Anaconda3 console and install git
#       > conda install GitPython
# After installing GitPython, run the application
# Open the localhost link http://127.0.0.1:5000/ to open the tool

# importing modules
from flask import Flask, request, redirect, url_for, render_template, session
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import git
from predict_funct import *

app = Flask(__name__)

# Secret key of "hello" used for sessions
app.secret_key = "hello"

# Pre-coded configurations
THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
app.config["EXCEL_UPLOADS"] = "static/assets/uploads"
app.config["ASSETS"] = "static/assets"
app.config["ALLOWED_EXCEL_EXTENSIONS"] = ["XLSX", "CSV", "XLS"]
my_excel = os.path.join(THIS_FOLDER, "static/assets/uploads")
my_assets = os.path.join(THIS_FOLDER, "static/assets")

# Functionality to host the app using github workflow (currently not working)
@app.route('/git_update', methods=['POST'])
def git_update():
    repo = git.Repo('./PCOS-Diagnosing-SVM-DT')
    origin = repo.remotes.origin
    repo.create_head('main',
                     origin.refs.main).set_tracking_branch(origin.refs.main).checkout()
    origin.pull()
    return '', 200

# Main route of the app
# The user will select which disease to classify
@app.route("/")
def home():
    return render_template("home-page.html")

# Index page for the Ovarian Cancer upon selecting the disease
@app.route("/ovarian_index")
def ovarian_index():
    return render_template("ovarian-select-page.html")

# Index page for the PCOS upon selecting the disease
@app.route("/pcos_index")
def pcos_index():
    return render_template("pcos-select-page.html")

# Function to accept excel files by their extension

def allowed_excel(filename): # filename contains the filename of the uploaded excel file
    if not "." in filename:
        return False

    ext = filename.rsplit(".", 1)[1]
    if ext.upper() in app.config["ALLOWED_EXCEL_EXTENSIONS"]: # If the excel file consist of the allowed extensions like .xlxs, .csv, .xlx
        return True
    else:
        return False

# Classifier page of the PCOS and when the user select the SVM classifier
@app.route("/pcos_svm", methods=["GET", "POST"])
def pcos_svm():

    # Method to delete any existing session data before classifying a new set of data
    session.pop("result", None)
    session.pop("model", None)

    if request.method == "POST": # POST method to get the uploaded excel file
        if request.files:
            excel = request.files["input"] # "input" is the excel which have been uploaded to the tool

            if excel.filename == "": # To check if the excel file has a filename
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename): # To check if the excel file have the allowed extension
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename) # To check if the excel file has a secured filename
                excel.save(os.path.join(my_excel, filename)) # Saving the uploaded excel file to the directory
                session['save_excel'] = filename # Saving the excel filename to a session

            output = predict_pcos_svm(excel) # Call function for predicting the uploaded excel file then saving the result to 'output'
            session['result'] = int(output) # Saving the results of 'output' to a session

            return redirect(url_for("pcos_results")) # Redirecting the page to the 'pcos_results' route
    else:
        if "result" in session: # Checking if a pre-existing session has been made, if yes, redirecting the page to the 'pop' route
            return redirect(url_for("pop"))
    return render_template("pcos-svm-page.html") # Rendering the template of the page

# Classifier page of the PCOS and when the user select the Decision Tree classifier
@app.route("/pcos_dt", methods=["GET", "POST"])
def pcos_dt():
    session.pop("result", None)
    session.pop("model", None)
    if request.method == "POST":
        if request.files:
            excel = request.files["input"]

            if excel.filename == "":
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename):
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename)
                excel.save(os.path.join(my_excel, filename))
                session['save_excel'] = filename

            output = predict_pcos_dt(excel)
            session['result'] = int(output)

            return redirect(url_for("pcos_results"))
    else:
        if "result" in session:
            return redirect(url_for("pop"))
    return render_template("pcos-dt-page.html")

# Classifier page of the Ovarian Cancer and when the user select the SVM classifier
@app.route("/ovarian_svm", methods=["GET", "POST"])
def ovarian_svm():
    session.pop("result", None)
    session.pop("model", None)

    if request.method == "POST":
        if request.files:
            excel = request.files["input"]

            if excel.filename == "":
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename):
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename)
                excel.save(os.path.join(my_excel, filename))
                session['save_excel'] = filename

            output = predict_ovarian_svm(excel)
            session['result'] = int(output)

            return redirect(url_for("ovarian_result"))
    else:
        if "result" in session:
            return redirect(url_for("pop"))
    return render_template("ovarian-svm-page.html")

# Classifier page of the Ovarian Cancer and when the user select the Decision Tree classifier
@app.route("/ovarian_dt", methods=["GET", "POST"])
def ovarian_dt():
    session.pop("result", None)
    session.pop("model", None)
    if request.method == "POST":
        if request.files:
            excel = request.files["input"]

            if excel.filename == "":
                print("Excel file must have a filename")
                return redirect(request.url)

            if not allowed_excel(excel.filename):
                print("That excel extension is not allowed")
                return redirect(request.url)

            else:
                filename = secure_filename(excel.filename)
                excel.save(os.path.join(my_excel, filename))
                session['save_excel'] = filename

            output = predict_ovarian_dt(excel)
            session['result'] = int(output)

            return redirect(url_for("ovarian_result"))
    else:
        if "result" in session:
            return redirect(url_for("pop"))
    return render_template("ovarian-dt-page.html")

# The results page of PCOS classifier
@app.route("/pcos_results", methods=["GET", "POST"]) # POST and GET method for getting the result and showing it
def pcos_results():
    save_excel = session['save_excel'] # Retrieving the excel file from the session
    book = load_workbook(open(os.path.join(my_excel, save_excel), 'rb')) # Loading the excel file
    sheet = book.active

    if "result" in session: # Checking if the 'result' session is in session, if yes, retrieving the sessions to be displayed on the page
        result = session["result"]
        model = session["model"]
        PatID = session["PatID"]
        Age = session["Age"]
        Hairgrowth = session["Hairgrowth"]
        CycleRI = session["CycleRI"]
        WeightGain = session["WeightGain"]
        FastFood = session["FastFood"]

        # Retrieving which model has been used to be displayed
        if model == "SVM":
            model_name = "SVM"
        else:
            model_name = "DT"

        print(result)

        # If the result is '1' then the result is positive
        if result == 1:
            return render_template("pcos-results-page.html", RESULTS="POSITIVE", EXCEL=sheet, MODEL=model_name, ID=PatID, AGE=Age,
                                   HAIR=Hairgrowth, CYC=CycleRI, WEG=WeightGain, FAF=FastFood) # Assigning variables to the render template to be displayed
        # If the result is '0' then the result is positive
        elif result == 0:
            return render_template("pcos-results-page.html", RESULTS="NEGATIVE", EXCEL=sheet, MODEL=model_name, ID=PatID, AGE=Age,
                                   HAIR=Hairgrowth, CYC=CycleRI, WEG=WeightGain, FAF=FastFood) # Assigning variables to the render template to be displayed
    else:
        return redirect(url_for("pcos_svm")) 

# The results page of Ovarian Cancer classifier
@app.route("/ovarian_result", methods=["GET", "POST"])
def ovarian_result():
    save_excel = session['save_excel']
    book = load_workbook(open(os.path.join(my_excel, save_excel), 'rb'))
    sheet = book.active

    if "result" in session:
        result = session["result"]
        model = session["model"]

        Menopause = session["Menopause"]
        CANine = session["CANine"]
        CASeven = session["CASeven"]
        AeFP = session["AeFP"]
        CAOneTwo = session["CAOneTwo"]
        Age = session["Age"]
        if model == "SVM":
            model_name = "SVM"
        else:
            model_name = "DT"

        print(result)
        if result == 0:
            return render_template("ovarian-results-page.html", RESULTS="POSITIVE", EXCEL=sheet, MODEL=model_name, AGE=Age,
                                   MENO=Menopause, CAN=CANine, CAS=CASeven, AFP=AeFP, CAT=CAOneTwo)
        elif result == 1:
            return render_template("ovarian-results-page.html", RESULTS="NEGATIVE", EXCEL=sheet, MODEL=model_name, AGE=Age,
                                   MENO=Menopause, CAN=CANine, CAS=CASeven, AFP=AeFP, CAT=CAOneTwo)
    else:
        return redirect(url_for("pcos_svm"))

# Route for the About Page
@app.route("/about_page")
def about_page():
    return render_template("about-page.html")

# Route for the pop method that will delete any existing sessions
@app.route("/pop")
def pop():
    session.pop("result", None)
    session.pop("model", None)
    session.pop("result", None)
    session.pop("model", None)
    session.pop("PatID", None)
    session.pop("Age", None)
    session.pop("Hairgrowth", None)
    session.pop("CycleRI", None)
    session.pop("WeightGain", None)
    session.pop("FastFood", None)
    return redirect(url_for("pcos_svm"))


if __name__ == "__main__":
    app.run(debug=True)
